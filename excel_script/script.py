from typing import Optional, Union
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook


class ExcelProcessor:
    def __init__(self, filepath: str, skip_rows: int = 1) -> None:
        """
        Инициализирует объект ExcelProcessor и загружает данные из файла.
        :param filepath: Путь к Excel файлу.
        :param skip_rows: Количество строк для пропуска перед чтением заголовков и данных.
        """
        self.filepath = filepath
        self.headers: list[str] = list()
        self.data_rows = list()
        self._load_data(to_skip=skip_rows)

    def _load_data(self, to_skip: int) -> None:
        """
        Загружает данные из Excel файла, пропуская первые to_skip строк
        Читает заголовок (или если строки заголовков то объединяет их),
        и загружает оставшиеся строки данных.
        :param to_skip: Количество строк для пропуска перед чтением заголовков и данных.
        """
        wb = load_workbook(self.filepath, read_only=True)
        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)

        for _ in range(to_skip):
            next(rows)

        header1, header2 = next(rows), next(rows)
        max_cols = max(len(header1), len(header2))

        self.headers = [
            f'{(header1[i] or '').strip()} {(header2[i] or '').strip()}'.strip()
            for i in range(max_cols)
        ]

        for row in rows:
            self.data_rows.append(row)

        wb.close()

    def filter_rows(self,
                    filter_col: str,
                    filter_val: str,
                    needed_cols: list[str]
                    ) -> list[list[Optional[Union[str, int, float]]]]:
        """
        Фильтрует строки данных, выбирая только те, где в столбце filter_col
        значение равно filter_val (игнорируя регистр), и возвращает только
        указанные столбцы needed_cols.

        :param filter_col: Название столба для фильтрации
        :param filter_val: Значение в столбце для фильтра
        :param needed_cols: Список столбцов, которые нужно вернуть
        """
        idx_map = [self.headers.index(col) for col in needed_cols]
        filter_idx = self.headers.index(filter_col)
        result = list()

        for row in self.data_rows:
            if not row[filter_idx]:
                continue
            if str(row[filter_idx]).strip().lower() == filter_val.lower():
                result.append([row[i] if i < len(row) else None for i in idx_map])

        return result

    @classmethod
    def save_filtered(cls,
                      rows: list[list[Optional[Union[str, int, float]]]],
                      out_file: str,
                      needed_cols: list[str]
                      ) -> None:
        """
        Сохраняет отфильтрованные строки в новый Excel файл
        :param rows: Список строк для сохранения.
        :param out_file: Путь для сохранения нового файла Excel.
        :param needed_cols: Список заголовков, которые нужно сохранить.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered Data"

        ws.append(needed_cols)

        for row in rows:
            ws.append(row)
        wb.save(out_file)


class ExcelFilterApp:
    def __init__(self, master: tk.Tk) -> None:
        """
        Инициализирует GUI-приложение для фильтрации данных из Excel-файлов.

        :param master: Главный объект Tkinter (корневое окно).
        """
        self.master = master
        master.title("Excel Filter")
        master.minsize(420, 380)

        style = ttk.Style()
        style.theme_use('clam')

        self.frm = ttk.Frame(master, padding=12)
        self.frm.pack(fill='both', expand=True)

        self.btn_open = ttk.Button(self.frm, text='Open Excel File', command=self.open_file)
        self.btn_open.grid(row=0, column=0, columnspan=2, pady=10, sticky='ew')

        ttk.Label(self.frm, text='Пропустить строк:').grid(row=1, column=0, sticky='e', pady=5)
        self.skip_var = tk.IntVar(value=8)
        self.spin_skip = ttk.Spinbox(self.frm, from_=0, to=50, textvariable=self.skip_var, width=5)
        self.spin_skip.grid(row=1, column=1, sticky='w', pady=5)

        ttk.Label(self.frm, text='Столбец для фильтра:').grid(row=2, column=0, pady=5, sticky='e')
        self.combo_col = ttk.Combobox(self.frm, state='readonly')
        self.combo_col.grid(row=2, column=1, pady=5, sticky='ew')

        ttk.Label(self.frm, text='Значение фильтра:').grid(row=3, column=0, pady=5, sticky='e')
        self.entry_val = ttk.Entry(self.frm)
        self.entry_val.grid(row=3, column=1, pady=5, sticky='ew')

        self.btn_save = ttk.Button(self.frm, text='Save Filtered File', command=self.save_file)
        self.btn_save.grid(row=4, column=0, columnspan=2, pady=15, sticky='ew')

        self.frm.columnconfigure(1, weight=1)
        self.needed_cols = ["ФИО", "Должность", "Отдел", "Дата найма", "Зарплата"]
        self.processor = None
        self.source_path = None

    def open_file(self) -> None:
        """
        Обработчик кнопки выбора Excel файла.
        Открывает диалог выбора файла, сохраняет путь и загружает данные с учетом пропуска строк.
        """
        file = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if not file:
            return
        self.source_path = file
        self.load_file_with_skip()

    def load_file_with_skip(self) -> None:
        """
        Загружает выбранный Excel файл с учетом заданного количества пропускаемых строк.
        Обновляет список столбцов для фильтрации.
        """
        if not self.source_path:
            messagebox.showwarning("Внимание", "Сначала выберите файл.")
            return

        skip = self.skip_var.get()
        try:
            self.processor = ExcelProcessor(self.source_path, skip_rows=skip)
            self.combo_col['values'] = self.processor.headers
            if self.processor.headers:
                self.combo_col.current(0)
            messagebox.showinfo(
                'Файл открыт',
                f'Доступные столбцы:\n{", ".join(self.processor.headers)}'
            )
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")

    def save_file(self) -> None:
        """
        Обработчик кнопки сохранения отфильтрованного Excel файла.
        Проверяет корректность введённых данных, фильтрует строки и сохраняет результат.
        """
        if not self.processor:
            messagebox.showwarning("Нет файла", "Сначала выберите исходный файл.")
            return

        col = self.combo_col.get().strip()
        val = self.entry_val.get().strip()

        if not col or not val:
            messagebox.showwarning("Ввод", "Введите имя столбца и значение для фильтрации.")
            return

        try:
            filtered = self.processor.filter_rows(col, val, self.needed_cols)
            save_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                title='Save',
            )
            if not save_path:
                return
            self.processor.save_filtered(filtered, save_path, self.needed_cols)
            messagebox.showinfo('Успех', f'Файл успешно сохранён {save_path}')
        except Exception as exp:
            messagebox.showinfo('Ошибка', f'Ошибка при сохранении {str(exp)}')


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelFilterApp(master=root)
    root.mainloop()
